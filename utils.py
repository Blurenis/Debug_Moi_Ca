# -----------------------------------------------------------------------------
# Imports and Dependency Management
# -----------------------------------------------------------------------------
import sys
import os
import json
import requests
import traceback
import re
import subprocess
import tempfile
from pathlib import Path
from typing import List, Dict, Optional, Set, Tuple, Union, Any

# GUI Framework
from PyQt6 import QtCore

# Optional Dependencies: Syntax Highlighting & Rendering
try:
    from pygments import highlight
    from pygments.lexers import get_lexer_by_name, TextLexer
    from pygments.formatters import HtmlFormatter
    PYGMENTS_AVAILABLE = True
except ImportError:
    PYGMENTS_AVAILABLE = False

try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

# Optional Dependencies: Document Parsing
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    import xlrd
except ImportError:
    openpyxl = xlrd = None


# -----------------------------------------------------------------------------
# OpenAI API Integration
# -----------------------------------------------------------------------------
def gpt4_1_request(messages: List[Dict[str, str]], model_name: Optional[str] = None) -> str:
    """
    Executes a standard chat completion request using the OpenAI API.

    Args:
        messages (List[Dict[str, str]]): A list of message dictionaries (roles and content).
        model_name (Optional[str]): The specific model ID to use. Defaults to "gpt-4o".

    Returns:
        str: The content of the model's response or an error message.
    """
    api_base: str = "https://api.openai.com/v1"
    
    # Securely retrieve API key from environment variables
    api_key: Optional[str] = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        return "API Error: The environment variable 'OPENAI_API_KEY' is not defined."

    default_model: str = "gpt-4o"
    target_model: str = model_name if model_name else default_model

    url: str = f"{api_base}/chat/completions"
    headers: Dict[str, str] = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    data: Dict[str, Any] = {
        "model": target_model,
        "messages": messages,
        "response_format": {"type": "text"}
    }
    
    try:
        response = requests.post(url, headers=headers, json=data, timeout=120)
        
        if response.status_code != 200:
            return f"API Error ({response.status_code}): {response.text}"
            
        response_json = response.json()
        return response_json['choices'][0]['message']['content']
        
    except Exception as e:
        return f"Request error: {e}\n{traceback.format_exc()}"


class GptWorker(QtCore.QThread):
    """
    Asynchronous worker thread for handling LLM API requests.
    Prevents the main GUI thread from freezing during network I/O.
    """
    finished = QtCore.pyqtSignal(str)
    
    def __init__(self, messages: List[Dict[str, str]], model_name: Optional[str] = None) -> None:
        super().__init__()
        self.messages: List[Dict[str, str]] = messages
        self.model_name: Optional[str] = model_name

    def run(self) -> None:
        result: str = gpt4_1_request(self.messages, self.model_name)
        self.finished.emit(result)


# -----------------------------------------------------------------------------
# Code Execution Logic
# -----------------------------------------------------------------------------
class CodeExecutionWorker(QtCore.QThread):
    """
    Worker thread that executes Python code in a separate subprocess.
    Captures standard output and standard error safely.
    """
    finished = QtCore.pyqtSignal(str, str)
    
    def __init__(self, code_to_execute: str) -> None:
        super().__init__()
        self.code_to_execute: str = code_to_execute

    def run(self) -> None:
        stdout: str = ""
        stderr: str = ""
        filepath: str = ""
        
        try:
            # Create a temporary file to hold the code execution context
            with tempfile.NamedTemporaryFile(suffix=".py", delete=False, mode='w', encoding='utf-8') as f:
                f.write(self.code_to_execute)
                filepath = f.name
            
            # Execute the script via the current system python interpreter
            process = subprocess.run(
                [sys.executable, filepath],
                capture_output=True, 
                text=True, 
                timeout=15,
                encoding='utf-8'
            )
            stdout = process.stdout
            stderr = process.stderr
            
        except subprocess.TimeoutExpired:
            stderr = "Execution Error: Code execution exceeded the 15-second timeout limit."
        except Exception as e:
            stderr = f"Subprocess error: {e}\n{traceback.format_exc()}"
        finally:
            # Cleanup temporary file
            if filepath and os.path.exists(filepath):
                os.unlink(filepath)
                
        self.finished.emit(stdout, stderr)


# -----------------------------------------------------------------------------
# File System & Content Extraction Logic
# -----------------------------------------------------------------------------
def human_readable_size(num: float) -> str:
    """Converts bytes to a human-readable string format (e.g., KB, MB)."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if num < 1024.0:
            return f"{num:3.1f} {unit}"
        num /= 1024.0
    return f"{num:.1f} PB"


class ProjectContextExtractor(QtCore.QObject):
    """
    Utilities for traversing directory structures and extracting file contents 
    based on configurable extension allowlists and directory blocklists.
    """
    
    DEFAULT_EXCLUSIONS: Set[str] = {
        "venv", "MyVenv", ".venv", "env", "log", "logs", ".env", "node_modules", ".git", "__pycache__",
        ".mypy_cache", ".pytest_cache", ".idea", ".vscode", ".DS_Store", ".cache",
        ".ipynb_checkpoints", ".history", ".svn", ".hg", ".tox", ".coverage",
        ".gitignore", ".gitattributes", ".yarn", ".parcel-cache", ".next", ".nuxt",
        ".node_modules", ".dist"
    }

    def __init__(self) -> None:
        super().__init__()
        self.exclusions: Set[str] = set()
        self.extensions: Set[str] = {
            '.py', '.pyw', '.ipynb', '.txt', '.pdf', '.docx', '.xlsx', '.xls',
            '.js', '.jsx', '.ts', '.tsx', '.vue', '.html', '.css',
            '.c', '.h', '.hpp', '.hh', '.cc', '.cpp', '.cxx', '.c++',
            '.sh', '.bash', '.zsh', '.ksh', '.bat', '.cmd', '.make', '.mk', 'Makefile',
            '.java', '.go', '.rs', '.swift', '.php', '.rb', '.pl', '.pm',
            '.scala', '.kt', '.kts', '.lua', '.sql',
            '.xml', '.yml', '.yaml', '.json', '.md', '.csv', '.ini', '.cfg', '.conf', '.toml',
            '.ps1', '.psm1', '.psd1', '.cmake', 'CMakeLists.txt', 'Dockerfile', 'dockerfile', 'Vagrantfile', 'Procfile'
        }

    def set_extensions(self, extension_list: List[str]) -> None:
        """Updates the list of file extensions to process."""
        self.extensions = set()
        for e in extension_list:
            e = e.strip().lower()
            if not e:
                continue
            if not e.startswith("."):
                e = "." + e
            self.extensions.add(e)

    def set_exclusions(self, exclusions: List[str]) -> None:
        """Updates the list of directories or files to ignore."""
        new_exclusions = set(x.strip() for x in exclusions if x.strip())
        self.exclusions = self.DEFAULT_EXCLUSIONS.union(new_exclusions)

    def build_context(self, folder_path: str, extract_content: bool = True) -> str:
        """
        Generates a comprehensive report of the folder structure and optionally the file contents.
        
        Args:
            folder_path (str): The root directory to scan.
            extract_content (bool): If True, appends the content of valid files.

        Returns:
            str: Formatted string containing structure and content.
        """
        self.exclusions = self.DEFAULT_EXCLUSIONS.union(self.exclusions)
        struct, cont, stats = self.get_folder_structure_and_content(
            folder_path, 
            extract_content=extract_content
        )
        
        result: List[str] = []
        result.append("FOLDER STRUCTURE:")
        result.extend(struct)
        
        if extract_content:
            result.append("\n######################################\n")
            result.append("FILE CONTENTS:")
            result.extend(cont)
            result.append("\n######################################\n")
            result.append(
                f"Statistics: {stats['words']} words, {stats['lines']} lines, "
                f"{stats['characters']} chars, ~{stats['tokens']} tokens. "
                f"Size: {human_readable_size(stats['size'])}"
            )
        return "\n".join(result).strip()

    def build_targeted_context(self, folder_path: str, target_files: List[str]) -> str:
        """
        Generates a context report containing the full directory structure,
        but includes content ONLY for the specified target files.
        """
        self.exclusions = self.DEFAULT_EXCLUSIONS.union(self.exclusions)
        
        # 1. Retrieve full folder structure (without content)
        struct, _, _ = self.get_folder_structure_and_content(folder_path, extract_content=False)
        
        content_output: List[str] = []
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0, 'size': 0}
        
        # 2. Extract content specific to the target files
        # Normalize paths for robust matching
        clean_targets: List[str] = [os.path.normpath(f).lower() for f in target_files]
        
        for root, dirs, files in os.walk(folder_path):
            # Apply directory exclusions
            dirs[:] = [d for d in dirs if d not in self.exclusions]
            
            for file in files:
                if file in self.exclusions: 
                    continue
                
                full_path: str = os.path.join(root, file)
                rel_path: str = os.path.relpath(full_path, folder_path)
                
                # Check if the file matches a target (by filename or relative path)
                if os.path.normpath(rel_path).lower() in clean_targets or \
                   file.lower() in clean_targets:
                    
                    file_content, file_stats = self.extract_file_content(full_path, indent_level=1)
                    content_output.append(f"\nFile: {rel_path}")
                    content_output.append(file_content)
                    
                    # Accumulate statistics
                    for k in stats:
                        if k in file_stats: 
                            stats[k] += file_stats[k]

        result: List[str] = []
        result.append("FOLDER STRUCTURE (Full):")
        result.extend(struct)
        result.append("\n######################################\n")
        result.append(f"SELECTED RELEVANT FILE CONTENTS ({len(content_output)//2} files):")
        result.extend(content_output)
        result.append("\n######################################\n")
        result.append(f"Targeted Stats: ~{stats['tokens']} tokens.")
        
        return "\n".join(result).strip()

    def get_folder_structure_and_content(self, folder_path: str, indent_level: int = 0, extract_content: bool = True) -> Tuple[List[str], List[str], Dict[str, int]]:
        """Recursively traverses the folder to build structure and content lists."""
        structure_output: List[str] = []
        content_output: List[str] = []
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0, 'size': 0}

        try:
            items: List[str] = sorted(os.listdir(folder_path))
        except Exception as ex:
            return [f"{'    ' * indent_level}[Cannot open: {ex}]"], [], stats

        for item in items:
            if item in self.exclusions:
                continue
                
            item_path: str = os.path.join(folder_path, item)
            
            if os.path.isdir(item_path):
                # Process Directory
                structure_output.append(f"{'    ' * indent_level}{item}/")
                sub_structure, sub_content, sub_stats = self.get_folder_structure_and_content(
                    item_path, 
                    indent_level + 1,
                    extract_content=extract_content
                )
                structure_output.extend(sub_structure)
                content_output.extend(sub_content)
                for key in stats:
                    stats[key] += sub_stats[key]
            else:
                # Process File
                ext: str = os.path.splitext(item_path)[1].lower()
                if ext not in self.extensions:
                    continue
                    
                structure_output.append(f"{'    ' * indent_level}{item}")
                
                if not extract_content:
                    continue
                    
                content_output.append(f"\n{'    ' * indent_level}File: {item_path}")
                try:
                    file_size = os.path.getsize(item_path)
                    stats['size'] += file_size
                    
                    # Skip large files to prevent performance issues
                    if file_size > 100 * 1024:
                        content_output.append(f"{'    ' * indent_level}[File too large to display]")
                    else:
                        file_content, file_stats = self.extract_file_content(item_path, indent_level)
                        content_output.append(file_content)
                        for key in stats:
                            if key in file_stats:
                                stats[key] += file_stats[key]
                except Exception as e:
                    content_output.append(f"{'    ' * indent_level}[Error reading file: {str(e)}]")
                    
        return structure_output, content_output, stats

    def extract_file_content(self, filepath: str, indent_level: int = 0) -> Tuple[str, Dict[str, int]]:
        """Determines file type and delegates extraction to the appropriate method."""
        ext: str = os.path.splitext(filepath)[1].lower()
        indent: str = '    ' * indent_level
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0}
        content: str = ""

        supported_text_exts: Set[str] = {
            '.txt', '.py', '.js', '.jsx', '.ts', '.tsx', '.vue',
            '.html', '.css', '.json', '.xml', '.md', '.csv', '.log',
            '.ini', '.cfg', '.conf', '.sh', '.bat', '.c', '.cpp', '.h',
            '.java', '.yml', '.yaml'
        }
        
        if ext in supported_text_exts:
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    file_content = f.read()
                content = f"{indent}Content:\n{file_content}"
                stats['words'] = len(file_content.split())
                stats['lines'] = file_content.count('\n')
                stats['characters'] = len(file_content)
                stats['tokens'] = len(file_content.split())
            except Exception as e:
                content = f"{indent}[Error reading text file: {e}]"
        
        elif ext in {'.xlsx', '.xls'}:
            content, stats = self.extract_excel_content(filepath, indent_level)
        elif ext == '.ipynb':
            content, stats = self.extract_ipynb_content(filepath, indent_level)
        elif ext == '.docx':
            if DOCX_AVAILABLE:
                content, stats = self.extract_docx_content(filepath, indent_level)
            else:
                content = f"{indent}[python-docx not installed]"
        elif ext == '.pdf':
            content = f"{indent}[PDF file (content not extracted)]"
        else:
            content = f"{indent}[Binary or unsupported file]"
            
        return content, stats

    def extract_excel_content(self, filepath: str, indent_level: int = 0) -> Tuple[str, Dict[str, int]]:
        """Extracts content from Excel files (xls/xlsx)."""
        ext: str = os.path.splitext(filepath)[1].lower()
        content_lines: List[str] = []
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0}
        indent: str = '    ' * indent_level
        
        try:
            if ext == '.xlsx' and openpyxl:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    content_lines.append(f"{indent}Sheet: {sheet}")
                    for row in ws.iter_rows(max_rows=50, max_cols=20, values_only=True):
                        row_str = "\t".join([str(cell)[:50] if cell is not None else "" for cell in row])
                        content_lines.append(f"{indent}{row_str}")
                        stats['lines'] += 1
                        stats['words'] += len(row_str.split())
                        stats['characters'] += len(row_str)
                        stats['tokens'] += len(row_str.split())
                        
            elif ext == '.xls' and xlrd:
                wb = xlrd.open_workbook(filepath)
                for sheet in wb.sheets():
                    content_lines.append(f"{indent}Sheet: {sheet.name}")
                    for row_idx in range(min(sheet.nrows, 50)):
                        row = sheet.row_values(row_idx)
                        row_str = "\t".join([str(cell)[:50] if cell is not None else "" for cell in row[:20]])
                        content_lines.append(f"{indent}{row_str}")
                        stats['lines'] += 1
                        stats['words'] += len(row_str.split())
                        stats['characters'] += len(row_str)
                        stats['tokens'] += len(row_str.split())
            else:
                content_lines.append(f"{indent}[Excel library not installed or unknown format]")
        except Exception as e:
            content_lines.append(f"{indent}[Error reading Excel file: {str(e)}]")
            
        return "\n".join(content_lines), stats

    def extract_ipynb_content(self, filepath: str, indent_level: int = 0) -> Tuple[str, Dict[str, int]]:
        """Extracts code and text from Jupyter Notebooks."""
        indent: str = '    ' * indent_level
        content_lines: List[str] = [f"{indent}[Jupyter Notebook]"]
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0}
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            cells = data.get('cells', [])
            
            for i, cell in enumerate(cells):
                cell_type = cell.get('cell_type', 'unknown')
                content_lines.append(f"{indent}Cell {i+1} [{cell_type}]:")
                source = cell.get('source', [])
                
                if isinstance(source, list):
                    cell_text = ''.join(source)
                else:
                    cell_text = str(source)
                    
                content_lines.append(f"{indent}{cell_text}")
                stats['lines'] += cell_text.count('\n') + 1
                stats['words'] += len(cell_text.split())
                stats['characters'] += len(cell_text)
                stats['tokens'] += len(cell_text.split())
        except Exception as e:
            content_lines.append(f"{indent}[Error reading ipynb: {str(e)}]")
            
        return "\n".join(content_lines), stats

    def extract_docx_content(self, filepath: str, indent_level: int = 0) -> Tuple[str, Dict[str, int]]:
        """Extracts text from Word documents."""
        indent: str = '    ' * indent_level
        content_lines: List[str] = [f"{indent}[Word Document]"]
        stats: Dict[str, int] = {'words': 0, 'lines': 0, 'characters': 0, 'tokens': 0}
        
        try:
            doc = Document(filepath)
            for para in doc.paragraphs:
                text = para.text
                content_lines.append(f"{indent}{text}")
                stats['lines'] += 1
                stats['words'] += len(text.split())
                stats['characters'] += len(text)
                stats['tokens'] += len(text.split())
        except Exception as e:
            content_lines.append(f"{indent}[Error reading docx: {str(e)}]")
            
        return "\n".join(content_lines), stats


# -----------------------------------------------------------------------------
# Markdown Rendering Logic
# -----------------------------------------------------------------------------
def markdown_to_html(md_text: str) -> str:
    """
    Converts Markdown text to HTML.
    Supports code block highlighting if Pygments is available.
    
    Args:
        md_text (str): The raw markdown string.
        
    Returns:
        str: The rendered HTML string.
    """
    if not MARKDOWN_AVAILABLE:
        return f"<pre>{md_text}</pre>"

    if PYGMENTS_AVAILABLE:
        from markdown.extensions import Extension
        from markdown.preprocessors import Preprocessor
        
        class CodeHiliteExtension(Extension):
            """Extension to register the custom CodeBlockHilite preprocessor."""
            def extendMarkdown(self, md: markdown.Markdown) -> None:
                md.registerExtension(self)
                md.preprocessors.register(CodeBlockHilite(), 'codehilite', 25)

        class CodeBlockHilite(Preprocessor):
            """Preprocessor to find code blocks and apply Pygments highlighting."""
            CODEBLOCK_RE = re.compile(r'```\s*(?P<lang>[\w+-]*)\n(?P<code>.*?)(?:```)', re.DOTALL)
            
            def run(self, lines: List[str]) -> List[str]:
                text = "\n".join(lines)
                
                def repl(match: re.Match) -> str:
                    code = match.group('code')
                    lang = match.group('lang') or 'text'
                    try:
                        lexer = get_lexer_by_name(lang) if lang else TextLexer()
                    except Exception:
                        lexer = TextLexer()
                    
                    formatter = HtmlFormatter(nowrap=False)
                    highlighted = highlight(code, lexer, formatter)
                    return f"\n<div class='hl_code'>{highlighted}</div>\n"
                    
                return self.CODEBLOCK_RE.sub(repl, text).splitlines()

        return markdown.markdown(
            md_text,
            extensions=['fenced_code', CodeHiliteExtension()],
            output_format='html5'
        )
    else:
        # Fallback if Pygments is not installed
        return markdown.markdown(
            md_text, 
            extensions=['fenced_code'], 
            output_format='html5'
        )